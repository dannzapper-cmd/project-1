"""In-memory file extraction adapters for Smart Lead Intake.

Block 10F-A deliberately keeps file handling as a thin adapter over the
existing preview/mapping/validation pipeline:

* CSV is decoded to text and sent through ``csv_text``.
* XLSX first-sheet rows are converted to ``records_json``.
* Text-based PDF tables are converted to ``records_json``.

No uploaded file is written to disk, executed, persisted, sent to a model, or
used for live research. Scanned/image-only PDFs are rejected with an explicit
OCR-needed message.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
import re
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import pdfplumber

from app.services.intake_normalizer import MAX_LEADS_PER_RUN

ExtractedInputType = Literal["csv_text", "records_json"]

OCR_REQUIRED_MESSAGE = (
    "This PDF appears to require OCR. Image/OCR intake is planned for the next block."
)
EMPTY_FIRST_SHEET_MESSAGE = (
    "The first sheet appears empty. Check that your lead table is on the first sheet."
)

_KNOWN_HEADER_NAMES: frozenset[str] = frozenset(
    {
        "account",
        "account_name",
        "category",
        "comments",
        "company",
        "company_name",
        "company name",
        "company_size",
        "contact",
        "contact_name",
        "contact name",
        "contact_role",
        "contact role",
        "country",
        "description",
        "domain",
        "employees",
        "employee_count",
        "full_name",
        "full name",
        "geography",
        "headcount",
        "id",
        "industry",
        "job_title",
        "job title",
        "lead_id",
        "lead id",
        "location",
        "market",
        "name",
        "notes",
        "organization",
        "position",
        "role",
        "sector",
        "site",
        "size",
        "title",
        "url",
        "vertical",
        "web",
        "website",
        "website_url",
    }
)

_REQUIRED_HEADER_ALIASES: dict[str, frozenset[str]] = {
    "company_name": frozenset(
        {
            "account",
            "account_name",
            "company",
            "company_name",
            "organization",
        }
    ),
    "industry": frozenset(
        {
            "category",
            "industry",
            "sector",
            "vertical",
        }
    ),
}


class IntakeFileExtractionError(Exception):
    """User-facing extraction failure."""

    def __init__(self, message: str, *, status_code: int = 422) -> None:
        super().__init__(message)
        self.message = message
        self.status_code = status_code


@dataclass(frozen=True)
class ExtractedFilePayload:
    """Rows ready for the existing intake preview pipeline."""

    input_type: ExtractedInputType
    detected_format: Literal["csv", "xlsx", "pdf"]
    content: str | None = None
    records: list[dict[str, Any]] | None = None
    found_rows: int = 0
    previewed_rows: int = 0
    warnings: list[str] = field(default_factory=list)


def extract_csv_text(file_bytes: bytes) -> ExtractedFilePayload:
    """Decode a UTF-8 CSV upload without trying alternate encodings."""

    try:
        content = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise IntakeFileExtractionError("File could not be decoded as UTF-8.") from exc

    if content.strip() == "":
        raise IntakeFileExtractionError("Uploaded file is empty.")

    return ExtractedFilePayload(
        input_type="csv_text",
        detected_format="csv",
        content=content,
    )


def extract_xlsx_records(file_bytes: bytes) -> ExtractedFilePayload:
    """Extract first-sheet XLSX rows as dictionaries keyed by detected headers."""

    try:
        workbook = load_workbook(
            filename=BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except (InvalidFileException, OSError, ValueError) as exc:
        raise IntakeFileExtractionError(
            "Excel file could not be read. Upload a valid .xlsx workbook."
        ) from exc

    worksheet = workbook.worksheets[0] if workbook.worksheets else None
    if worksheet is None:
        raise IntakeFileExtractionError(EMPTY_FIRST_SHEET_MESSAGE)

    rows = [_trim_row(row) for row in worksheet.iter_rows(values_only=True)]
    non_empty_rows = [row for row in rows if _row_has_values(row)]
    if len(non_empty_rows) < 2:
        raise IntakeFileExtractionError(EMPTY_FIRST_SHEET_MESSAGE)

    header_index = _detect_header_row_index(non_empty_rows)
    warnings: list[str] = []
    if header_index > 0:
        warnings.append(
            f"Detected the Excel header on row {header_index + 1} of the first sheet."
        )
    elif not _row_has_known_header(non_empty_rows[0]):
        warnings.append(
            "Excel headers were not recognized with high confidence; review the detected mapping."
        )

    headers = [_header_name(value, idx) for idx, value in enumerate(non_empty_rows[header_index])]
    data_rows = non_empty_rows[header_index + 1 :]
    if not data_rows:
        raise IntakeFileExtractionError(EMPTY_FIRST_SHEET_MESSAGE)

    records = _rows_to_records(headers, data_rows)
    records, found_rows, limit_warnings = _limit_records(records, "Excel")
    warnings.extend(limit_warnings)
    warnings.extend(_missing_required_column_warnings(headers, "Excel"))

    if not records:
        raise IntakeFileExtractionError("Excel extraction returned no lead rows.")

    return ExtractedFilePayload(
        input_type="records_json",
        detected_format="xlsx",
        records=records,
        found_rows=found_rows,
        previewed_rows=len(records),
        warnings=warnings,
    )


def extract_pdf_records(file_bytes: bytes) -> ExtractedFilePayload:
    """Extract simple text-based PDF table rows.

    The preferred path uses pdfplumber table extraction. A conservative text
    fallback handles PDFs where table-like columns are separated by tabs,
    commas, or repeated spaces. Image-only/scanned PDFs are rejected.
    """

    try:
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            text_by_page = [page.extract_text() or "" for page in pdf.pages]
            if not any(text.strip() for text in text_by_page):
                raise IntakeFileExtractionError(OCR_REQUIRED_MESSAGE)

            table_rows: list[list[Any]] = []
            for page in pdf.pages:
                for table in page.extract_tables() or []:
                    for row in table:
                        trimmed = _trim_row(row)
                        if _row_has_values(trimmed):
                            table_rows.append(trimmed)

            warnings: list[str] = []
            if table_rows:
                records, header_warning = _records_from_table_rows(table_rows)
                if header_warning:
                    warnings.append(header_warning)
            else:
                records = _records_from_pdf_text_lines(text_by_page)
                warnings.append(
                    "PDF text extraction confidence is limited; review the preview carefully."
                )
    except IntakeFileExtractionError:
        raise
    except Exception as exc:
        raise IntakeFileExtractionError(
            "PDF file could not be read. Upload a text-based PDF table."
        ) from exc

    if not records:
        raise IntakeFileExtractionError(
            "No table-like rows could be extracted from this PDF. PDF extraction works best with text-based tables."
        )

    headers = list(records[0].keys()) if records else []
    records, found_rows, limit_warnings = _limit_records(records, "PDF")
    warnings.extend(limit_warnings)
    warnings.extend(_missing_required_column_warnings(headers, "PDF"))

    return ExtractedFilePayload(
        input_type="records_json",
        detected_format="pdf",
        records=records,
        found_rows=found_rows,
        previewed_rows=len(records),
        warnings=warnings,
    )


def _records_from_table_rows(table_rows: list[list[Any]]) -> tuple[list[dict[str, Any]], str | None]:
    header_index = _detect_header_row_index(table_rows)
    header_warning: str | None = None
    if not _row_has_known_header(table_rows[header_index]):
        header_warning = (
            "PDF table headers were not recognized with high confidence; review the detected mapping."
        )
    elif header_index > 0:
        header_warning = f"Detected the PDF table header on extracted row {header_index + 1}."

    headers = [_header_name(value, idx) for idx, value in enumerate(table_rows[header_index])]
    return _rows_to_records(headers, table_rows[header_index + 1 :]), header_warning


def _records_from_pdf_text_lines(text_by_page: list[str]) -> list[dict[str, Any]]:
    lines = [
        line.strip()
        for page_text in text_by_page
        for line in page_text.splitlines()
        if line.strip()
    ]
    split_lines = [_split_table_line(line) for line in lines]
    candidate_indexes = [
        idx
        for idx, row in enumerate(split_lines)
        if len(row) >= 2 and _row_has_known_header(row)
    ]
    if not candidate_indexes:
        return []

    header_index = candidate_indexes[0]
    headers = [_header_name(value, idx) for idx, value in enumerate(split_lines[header_index])]
    data_rows: list[list[Any]] = []
    for row in split_lines[header_index + 1 :]:
        if len(row) < 2:
            continue
        data_rows.append(row)

    return _rows_to_records(headers, data_rows)


def _split_table_line(line: str) -> list[str]:
    if "\t" in line:
        return [cell.strip() for cell in line.split("\t") if cell.strip()]
    if line.count(",") >= 2:
        return [cell.strip() for cell in line.split(",") if cell.strip()]
    return [cell.strip() for cell in re.split(r"\s{2,}", line) if cell.strip()]


def _trim_row(row: tuple[Any, ...] | list[Any]) -> list[Any]:
    trimmed: list[Any] = []
    for value in row:
        if isinstance(value, str):
            trimmed.append(value.strip())
        else:
            trimmed.append(value)
    return trimmed


def _row_has_values(row: list[Any]) -> bool:
    return any(_cell_to_string(value) != "" for value in row)


def _detect_header_row_index(rows: list[list[Any]]) -> int:
    for idx, row in enumerate(rows[:10]):
        non_empty = [_cell_to_string(value) for value in row if _cell_to_string(value)]
        if len(non_empty) >= 2 and _row_has_known_header(row):
            return idx
    return 0


def _row_has_known_header(row: list[Any]) -> bool:
    normalized = {_normalize_header(_cell_to_string(value)) for value in row}
    return any(value in _KNOWN_HEADER_NAMES for value in normalized if value)


def _rows_to_records(headers: list[str], rows: list[list[Any]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in rows:
        if not _row_has_values(row):
            continue
        record: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            value = row[idx] if idx < len(row) else None
            if isinstance(value, str):
                value = value.strip()
            record[header] = value
        records.append(record)
    return records


def _limit_records(
    records: list[dict[str, Any]],
    format_label: str,
) -> tuple[list[dict[str, Any]], int, list[str]]:
    found_rows = len(records)
    if found_rows <= MAX_LEADS_PER_RUN:
        return records, found_rows, []
    return (
        records[:MAX_LEADS_PER_RUN],
        found_rows,
        [
            (
                f"{format_label} extraction found {found_rows} rows; "
                f"previewing the first {MAX_LEADS_PER_RUN} rows to match the intake limit."
            )
        ],
    )


def _missing_required_column_warnings(headers: list[str], format_label: str) -> list[str]:
    normalized_headers = {_normalize_header(header) for header in headers}
    missing = sorted(
        field
        for field, aliases in _REQUIRED_HEADER_ALIASES.items()
        if normalized_headers.isdisjoint(aliases)
    )
    if not missing:
        return []
    return [
        (
            f"{format_label} extraction did not find required column(s): "
            f"{', '.join(missing)}. Rows missing company_name or industry cannot be processed."
        )
    ]


def _header_name(value: Any, idx: int) -> str:
    header = _cell_to_string(value)
    return header if header else f"Column {idx + 1}"


def _cell_to_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value: str) -> str:
    return value.strip().lower().replace(" ", "_").replace("-", "_")


__all__ = [
    "EMPTY_FIRST_SHEET_MESSAGE",
    "OCR_REQUIRED_MESSAGE",
    "ExtractedFilePayload",
    "IntakeFileExtractionError",
    "extract_csv_text",
    "extract_pdf_records",
    "extract_xlsx_records",
]
